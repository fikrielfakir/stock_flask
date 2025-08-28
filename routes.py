from flask import jsonify, request, send_file, make_response
from datetime import datetime
import uuid
import time
from sqlalchemy import or_, func, desc, and_
import pandas as pd
import io
import csv
from werkzeug.utils import secure_filename

def register_routes(app, db):
    from flask_models import Article, Supplier, Requestor, PurchaseRequest, PurchaseRequestItem, Reception, Outbound, ActivityLog, User, UserSession
    import logging
    import json
    logger = logging.getLogger(__name__)
    
    # Activity logging helper functions
    def log_activity(action, entity_type, entity_id=None, entity_name=None, old_values=None, new_values=None):
        """Log user activity to the database"""
        try:
            activity_log = ActivityLog(
                action=action,
                entity_type=entity_type,
                entity_id=entity_id,
                entity_name=entity_name,
                old_values=json.dumps(old_values) if old_values else None,
                new_values=json.dumps(new_values) if new_values else None,
                ip_address=request.remote_addr if request else None,
                user_agent=request.headers.get('User-Agent') if request else None
            )
            db.session.add(activity_log)
            db.session.commit()
        except Exception as e:
            logger.error(f"Failed to log activity: {str(e)}")
            # Don't fail the main operation if logging fails
            pass
    
    def get_entity_name(entity_type, entity_data):
        """Generate a human-readable name for the entity"""
        if entity_type == 'suppliers':
            return entity_data.get('nom', 'Fournisseur inconnu')
        elif entity_type == 'requestors':
            prenom = entity_data.get('prenom', '')
            nom = entity_data.get('nom', '')
            return f"{prenom} {nom}".strip() or 'Demandeur inconnu'
        elif entity_type == 'articles':
            return entity_data.get('designation', 'Article inconnu')
        return str(entity_data.get('id', 'Inconnu'))
    
    # Load settings at startup  
    def load_system_settings():
        try:
            import json
            import os
            if os.path.exists('settings.json'):
                with open('settings.json', 'r', encoding='utf-8') as f:
                    app.config['SYSTEM_SETTINGS'] = json.load(f)
            else:
                app.config['SYSTEM_SETTINGS'] = {}
        except Exception as e:
            app.logger.warning(f'Could not load system settings: {e}')
    
    # Load settings when routes are registered
    load_system_settings()
    
    # Authentication API endpoints
    @app.route("/api/auth/login", methods=['POST'])
    def login():
        try:
            data = request.get_json()
            username = data.get('username')
            password = data.get('password')
            
            if not username or not password:
                return jsonify({'message': 'Nom d\'utilisateur et mot de passe requis'}), 400
            
            # Find user
            user = User.query.filter_by(username=username).first()
            
            if not user or not user.check_password(password):
                return jsonify({'message': 'Nom d\'utilisateur ou mot de passe incorrect'}), 401
            
            if not user.is_active:
                return jsonify({'message': 'Compte désactivé'}), 401
            
            # Create new session (24-hour duration)
            session = UserSession.create_session(user.id)
            db.session.add(session)
            
            # Update last login
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            # Log login activity
            log_activity('LOGIN', 'users', user.id, user.username)
            
            return jsonify({
                'message': 'Connexion réussie',
                'sessionToken': session.session_token,
                'expiresAt': session.expires_at.isoformat(),
                'user': user.to_dict()
            })
            
        except Exception as e:
            logger.error(f"Login error: {str(e)}")
            return jsonify({'message': 'Erreur interne du serveur'}), 500
    
    @app.route("/api/auth/logout", methods=['POST'])
    def logout():
        try:
            auth_header = request.headers.get('Authorization')
            if auth_header and auth_header.startswith('Bearer '):
                token = auth_header.split(' ')[1]
                session = UserSession.query.filter_by(session_token=token).first()
                
                if session:
                    # Log logout activity
                    log_activity('LOGOUT', 'users', session.user_id, session.user.username)
                    
                    # Delete session
                    db.session.delete(session)
                    db.session.commit()
            
            return jsonify({'message': 'Déconnexion réussie'})
            
        except Exception as e:
            logger.error(f"Logout error: {str(e)}")
            return jsonify({'message': 'Erreur lors de la déconnexion'}), 500
    
    @app.route("/api/auth/verify", methods=['GET'])
    def verify_session():
        try:
            auth_header = request.headers.get('Authorization')
            if not auth_header or not auth_header.startswith('Bearer '):
                return jsonify({'message': 'Token manquant'}), 401
            
            token = auth_header.split(' ')[1]
            session = UserSession.query.filter_by(session_token=token).first()
            
            if not session or not session.is_valid():
                return jsonify({'message': 'Session expirée'}), 401
            
            # Update last activity
            session.user.last_login = datetime.utcnow()
            db.session.commit()
            
            return jsonify({
                'valid': True,
                'user': session.user.to_dict(),
                'expiresAt': session.expires_at.isoformat()
            })
            
        except Exception as e:
            logger.error(f"Session verification error: {str(e)}")
            return jsonify({'message': 'Erreur de vérification'}), 500
    
    # Smart UX API endpoints
    @app.route("/api/notifications", methods=['GET'])
    def get_notifications():
        try:
            # Real-time notifications
            notifications = []
            
            # Check low stock articles
            low_stock = Article.query.filter(Article.stock_actuel <= Article.stock_minimum).all()
            if low_stock:
                notifications.append({
                    'id': 'low_stock',
                    'type': 'warning',
                    'title': 'Stock critique',
                    'message': f'{len(low_stock)} articles en rupture de stock',
                    'time': 'Il y a 15 minutes',
                    'icon': 'fas fa-exclamation',
                    'color': 'red'
                })
            
            # Check pending requests
            pending_requests = PurchaseRequest.query.filter_by(statut='En attente').count()
            if pending_requests > 0:
                notifications.append({
                    'id': 'pending_requests',
                    'type': 'info',
                    'title': 'Demandes en attente',
                    'message': f'{pending_requests} demandes nécessitent votre attention',
                    'time': 'Il y a 30 minutes',
                    'icon': 'fas fa-shopping-cart',
                    'color': 'blue'
                })
            
            # Check recent receptions
            from datetime import datetime, timedelta
            recent_receptions = Reception.query.filter(
                Reception.date_reception >= datetime.now() - timedelta(hours=24)
            ).count()
            
            if recent_receptions > 0:
                notifications.append({
                    'id': 'recent_receptions',
                    'type': 'success',
                    'title': 'Nouvelles réceptions',
                    'message': f'{recent_receptions} livraisons reçues aujourd\'hui',
                    'time': 'Il y a 2 heures',
                    'icon': 'fas fa-truck',
                    'color': 'green'
                })
            
            # Add recent activities from activity logs
            recent_activities = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(5).all()
            for activity in recent_activities:
                # Calculate time ago
                time_diff = datetime.now() - activity.created_at
                if time_diff.seconds < 3600:
                    time_ago = f'Il y a {time_diff.seconds // 60} minutes'
                elif time_diff.days == 0:
                    time_ago = f'Il y a {time_diff.seconds // 3600} heures'
                else:
                    time_ago = f'Il y a {time_diff.days} jours'
                
                # Get icon based on action
                action_icons = {
                    'CREATE': 'fas fa-plus-circle',
                    'UPDATE': 'fas fa-edit',
                    'DELETE': 'fas fa-trash',
                    'EXPORT': 'fas fa-download',
                    'IMPORT': 'fas fa-upload'
                }
                
                notifications.append({
                    'id': f'activity_{activity.id}',
                    'type': 'info',
                    'title': 'Activité récente',
                    'message': activity.get_description(),
                    'time': time_ago,
                    'icon': action_icons.get(activity.action, 'fas fa-info-circle'),
                    'color': 'blue'
                })
                
            return jsonify({
                'notifications': notifications,
                'count': len(notifications)
            })
        except Exception as e:
            return jsonify({'notifications': [], 'count': 0, 'error': str(e)})

    # Activity Logs API
    @app.route("/api/activity-logs", methods=['GET'])
    def get_activity_logs():
        try:
            page = request.args.get('page', 1, type=int)
            limit = request.args.get('limit', 50, type=int)
            entity_type = request.args.get('entity_type')
            action = request.args.get('action')
            
            # Build query
            query = ActivityLog.query
            
            if entity_type:
                query = query.filter(ActivityLog.entity_type == entity_type)
            if action:
                query = query.filter(ActivityLog.action == action)
            
            # Paginate and order by newest first
            logs = query.order_by(ActivityLog.created_at.desc()).limit(limit).offset((page - 1) * limit).all()
            total_count = query.count()
            
            return jsonify({
                'logs': [log.to_dict() for log in logs],
                'totalCount': total_count,
                'page': page,
                'limit': limit
            })
        except Exception as e:
            logger.error(f"Activity logs error: {str(e)}")
            return jsonify({'message': f'Erreur lors de la récupération des logs: {str(e)}'}), 500
    
    @app.route("/api/quick-stats", methods=['GET'])
    def get_quick_stats():
        try:
            stats = {
                'total_articles': Article.query.count(),
                'low_stock_count': Article.query.filter(Article.stock_actuel <= Article.stock_minimum).count(),
                'pending_requests': PurchaseRequest.query.filter_by(statut='En attente').count(),
                'total_suppliers': Supplier.query.count(),
                'recent_activity': {
                    'new_articles_today': 0,  # Would calculate based on creation date
                    'processed_requests_today': 0,
                    'received_items_today': 0
                }
            }
            return jsonify(stats)
        except Exception as e:
            return jsonify({'error': str(e)})
    
    # Articles routes
    @app.route("/api/articles", methods=['GET'])
    def get_articles():
        try:
            # Get pagination parameters
            page = request.args.get('page', 1, type=int)
            per_page = request.args.get('per_page', 20, type=int)  # Default 20 per page
            
            # Get search and filter parameters
            search = request.args.get('search', '', type=str)
            category = request.args.get('category', 'all', type=str)
            stock_filter = request.args.get('stock_filter', 'all', type=str)
            
            # Build query with filters
            query = Article.query
            
            # Apply search filter
            if search:
                search_term = f"%{search}%"
                query = query.filter(
                    or_(
                        Article.designation.ilike(search_term),
                        Article.code_article.ilike(search_term),
                        Article.reference.ilike(search_term),
                        Article.marque.ilike(search_term)
                    )
                )
            
            # Apply category filter
            if category != 'all':
                query = query.filter(Article.categorie == category)
            
            # Apply stock level filter
            if stock_filter == 'low':
                query = query.filter(Article.stock_actuel <= Article.seuil_minimum)
            elif stock_filter == 'normal':
                query = query.filter(
                    and_(
                        Article.stock_actuel > Article.seuil_minimum,
                        Article.stock_actuel <= Article.seuil_minimum * 3
                    )
                )
            elif stock_filter == 'high':
                query = query.filter(Article.stock_actuel > Article.seuil_minimum * 3)
            
            # Get paginated results
            pagination = query.paginate(
                page=page,
                per_page=per_page,
                error_out=False
            )
            
            articles = [article.to_dict() for article in pagination.items]
            
            return jsonify({
                'articles': articles,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': pagination.total,
                    'pages': pagination.pages,
                    'has_prev': pagination.has_prev,
                    'has_next': pagination.has_next,
                    'prev_num': pagination.prev_num,
                    'next_num': pagination.next_num
                }
            })
        except Exception as e:
            logger.error(f"Error getting articles: {str(e)}")
            return jsonify({'message': 'Erreur lors de la récupération des articles'}), 500

    @app.route("/api/articles/search", methods=['GET'])
    def search_articles():
        try:
            query = request.args.get('query', '')
            if not query or len(query) < 3:
                return jsonify([])
            
            search_term = f'%{query.lower()}%'
            articles = Article.query.filter(
                or_(
                    func.lower(Article.designation).like(search_term),
                    func.lower(Article.code_article).like(search_term),
                    func.lower(Article.reference).like(search_term),
                    func.lower(Article.categorie).like(search_term)
                )
            ).limit(10).all()
            
            return jsonify([article.to_dict() for article in articles])
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la recherche d\'articles'}), 500

    @app.route("/api/search/global", methods=['GET'])
    def global_search():
        try:
            query = request.args.get('query', '')
            if not query or len(query) < 2:
                return jsonify({'results': [], 'totalCount': 0})
            
            search_term = f'%{query.lower()}%'
            results = []
            
            # Search Articles
            articles = Article.query.filter(
                or_(
                    func.lower(Article.designation).like(search_term),
                    func.lower(Article.code_article).like(search_term),
                    func.lower(Article.reference).like(search_term),
                    func.lower(Article.categorie).like(search_term)
                )
            ).limit(5).all()
            
            for article in articles:
                results.append({
                    'type': 'article',
                    'id': article.id,
                    'title': article.designation,
                    'subtitle': f'{article.code_article} - {article.categorie}',
                    'extra': f'Stock: {article.stock_actuel}',
                    'path': '/articles',
                    'data': article.to_dict()
                })
            
            # Search Suppliers
            suppliers = Supplier.query.filter(
                or_(
                    func.lower(Supplier.nom).like(search_term),
                    func.lower(Supplier.contact).like(search_term)
                )
            ).limit(5).all()
            
            for supplier in suppliers:
                results.append({
                    'type': 'supplier',
                    'id': supplier.id,
                    'title': supplier.nom,
                    'subtitle': supplier.contact or 'Pas de contact',
                    'extra': supplier.adresse or '',
                    'path': '/suppliers',
                    'data': supplier.to_dict()
                })
            
            # Search Requestors
            requestors = Requestor.query.filter(
                or_(
                    func.lower(func.concat(Requestor.prenom, ' ', Requestor.nom)).like(search_term),
                    func.lower(Requestor.departement).like(search_term),
                    func.lower(Requestor.poste).like(search_term)
                )
            ).limit(5).all()
            
            for requestor in requestors:
                results.append({
                    'type': 'requestor',
                    'id': requestor.id,
                    'title': f'{requestor.prenom} {requestor.nom}',
                    'subtitle': requestor.departement,
                    'extra': requestor.poste or '',
                    'path': '/requestors',
                    'data': requestor.to_dict()
                })
            
            return jsonify({
                'results': results[:15],
                'totalCount': len(results),
                'categories': {
                    'articles': len([r for r in results if r['type'] == 'article']),
                    'suppliers': len([r for r in results if r['type'] == 'supplier']),
                    'requestors': len([r for r in results if r['type'] == 'requestor'])
                }
            })
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la recherche globale'}), 500

    @app.route("/api/articles/<article_id>", methods=['GET'])
    def get_article(article_id):
        try:
            article = Article.query.get(article_id)
            if not article:
                return jsonify({'message': 'Article non trouvé'}), 404
            return jsonify(article.to_dict())
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération de l\'article'}), 500

    @app.route("/api/articles", methods=['POST'])
    def create_article():
        try:
            data = request.get_json()
            article = Article(
                code_article=data['codeArticle'],
                designation=data['designation'],
                categorie=data['categorie'],
                marque=data.get('marque'),
                reference=data.get('reference'),
                stock_initial=data.get('stockInitial', 0),
                stock_actuel=data.get('stockInitial', 0),
                unite=data.get('unite', 'pcs'),
                prix_unitaire=data.get('prixUnitaire'),
                seuil_minimum=data.get('seuilMinimum', 10),
                fournisseur_id=data.get('fournisseurId')
            )
            db.session.add(article)
            db.session.commit()
            return jsonify(article.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Données invalides', 'error': str(e)}), 400

    @app.route("/api/articles/<article_id>", methods=['PUT'])
    def update_article(article_id):
        try:
            article = Article.query.get(article_id)
            if not article:
                return jsonify({'message': 'Article non trouvé'}), 404
            
            data = request.get_json()
            for key, value in data.items():
                if hasattr(article, key.replace('C', '_c').replace('A', '_a').replace('I', '_i').replace('M', '_m').replace('S', '_s').replace('P', '_p').replace('F', '_f').lower()):
                    setattr(article, key.replace('C', '_c').replace('A', '_a').replace('I', '_i').replace('M', '_m').replace('S', '_s').replace('P', '_p').replace('F', '_f').lower(), value)
            
            db.session.commit()
            return jsonify(article.to_dict())
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erreur lors de la mise à jour', 'error': str(e)}), 400

    @app.route("/api/articles/<article_id>", methods=['DELETE'])
    def delete_article(article_id):
        try:
            article = Article.query.get(article_id)
            if not article:
                return jsonify({'message': 'Article non trouvé'}), 404
            
            db.session.delete(article)
            db.session.commit()
            return '', 204
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erreur lors de la suppression'}), 500

    @app.route("/api/articles/bulk-delete", methods=['POST'])
    def bulk_delete_articles():
        try:
            data = request.get_json()
            article_ids = data.get('ids', [])
            
            if not article_ids:
                return jsonify({'message': 'Aucun ID fourni'}), 400
            
            # Find articles to delete
            articles = Article.query.filter(Article.id.in_(article_ids)).all()
            deleted_count = len(articles)
            
            # Delete articles
            for article in articles:
                db.session.delete(article)
            
            db.session.commit()
            
            # Log activity
            log_activity(
                action='BULK_DELETE',
                entity_type='articles',
                entity_name=f'{deleted_count} articles',
                old_values={'count': deleted_count, 'ids': article_ids}
            )
            
            return jsonify({
                'message': f'{deleted_count} articles supprimés avec succès',
                'deletedCount': deleted_count
            }), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Bulk delete articles error: {str(e)}")
            return jsonify({'message': 'Erreur lors de la suppression en masse'}), 500

    @app.route("/api/articles/low-stock", methods=['GET'])
    def get_low_stock_articles():
        try:
            articles = Article.query.filter(Article.stock_actuel <= Article.seuil_minimum).all()
            return jsonify([article.to_dict() for article in articles])
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération des articles en stock bas'}), 500

    # Suppliers routes
    @app.route("/api/suppliers", methods=['GET'])
    def get_suppliers():
        try:
            suppliers = Supplier.query.all()
            return jsonify([supplier.to_dict() for supplier in suppliers])
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération des fournisseurs'}), 500

    @app.route("/api/suppliers", methods=['POST'])
    def create_supplier():
        try:
            data = request.get_json()
            supplier = Supplier(
                nom=data['nom'],
                contact=data.get('contact'),
                telephone=data.get('telephone'),
                email=data.get('email'),
                adresse=data.get('adresse'),
                conditions_paiement=data.get('conditionsPaiement'),
                delai_livraison=data.get('delaiLivraison')
            )
            db.session.add(supplier)
            db.session.commit()
            
            # Log activity
            log_activity(
                action='CREATE',
                entity_type='suppliers',
                entity_id=supplier.id,
                entity_name=get_entity_name('suppliers', data),
                new_values=data
            )
            
            return jsonify(supplier.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Données invalides', 'error': str(e)}), 400

    @app.route("/api/suppliers/<supplier_id>", methods=['PUT'])
    def update_supplier(supplier_id):
        try:
            supplier = Supplier.query.get(supplier_id)
            if not supplier:
                return jsonify({'message': 'Fournisseur non trouvé'}), 404
            
            data = request.get_json()
            
            # Store old values for logging
            old_values = supplier.to_dict()
            
            for key, value in data.items():
                snake_key = key.replace('conditionsPaiement', 'conditions_paiement').replace('delaiLivraison', 'delai_livraison')
                if hasattr(supplier, snake_key):
                    setattr(supplier, snake_key, value)
            
            db.session.commit()
            
            # Log activity
            log_activity(
                action='UPDATE',
                entity_type='suppliers',
                entity_id=supplier.id,
                entity_name=get_entity_name('suppliers', data),
                old_values=old_values,
                new_values=data
            )
            
            return jsonify(supplier.to_dict())
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erreur lors de la mise à jour', 'error': str(e)}), 400

    @app.route("/api/suppliers/<supplier_id>", methods=['DELETE'])
    def delete_supplier(supplier_id):
        try:
            supplier = Supplier.query.get(supplier_id)
            if not supplier:
                return jsonify({'message': 'Fournisseur non trouvé'}), 404
            
            # Store values for logging before deletion
            old_values = supplier.to_dict()
            entity_name = get_entity_name('suppliers', old_values)
            
            db.session.delete(supplier)
            db.session.commit()
            
            # Log activity
            log_activity(
                action='DELETE',
                entity_type='suppliers',
                entity_id=supplier_id,
                entity_name=entity_name,
                old_values=old_values
            )
            
            return '', 204
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erreur lors de la suppression'}), 500

    @app.route("/api/suppliers/bulk-delete", methods=['POST'])
    def bulk_delete_suppliers():
        try:
            data = request.get_json()
            supplier_ids = data.get('ids', [])
            
            if not supplier_ids:
                return jsonify({'message': 'Aucun ID fourni'}), 400
            
            # Find suppliers to delete
            suppliers = Supplier.query.filter(Supplier.id.in_(supplier_ids)).all()
            deleted_count = len(suppliers)
            
            # Delete suppliers
            for supplier in suppliers:
                db.session.delete(supplier)
            
            db.session.commit()
            
            # Log activity
            log_activity(
                action='BULK_DELETE',
                entity_type='suppliers',
                entity_name=f'{deleted_count} fournisseurs',
                old_values={'count': deleted_count, 'ids': supplier_ids}
            )
            
            return jsonify({
                'message': f'{deleted_count} fournisseurs supprimés avec succès',
                'deletedCount': deleted_count
            }), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Bulk delete suppliers error: {str(e)}")
            return jsonify({'message': 'Erreur lors de la suppression en masse'}), 500

    # Requestors routes
    @app.route("/api/requestors", methods=['GET'])
    def get_requestors():
        try:
            requestors = Requestor.query.all()
            return jsonify([requestor.to_dict() for requestor in requestors])
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération des demandeurs'}), 500

    @app.route("/api/requestors", methods=['POST'])
    def create_requestor():
        try:
            data = request.get_json()
            requestor = Requestor(
                nom=data['nom'],
                prenom=data['prenom'],
                departement=data['departement'],
                poste=data.get('poste'),
                email=data.get('email'),
                telephone=data.get('telephone')
            )
            db.session.add(requestor)
            db.session.commit()
            
            # Log activity
            log_activity(
                action='CREATE',
                entity_type='requestors',
                entity_id=requestor.id,
                entity_name=get_entity_name('requestors', data),
                new_values=data
            )
            
            return jsonify(requestor.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Données invalides', 'error': str(e)}), 400

    @app.route("/api/requestors/<requestor_id>", methods=['PUT'])
    def update_requestor(requestor_id):
        try:
            requestor = Requestor.query.get(requestor_id)
            if not requestor:
                return jsonify({'message': 'Demandeur non trouvé'}), 404
            
            data = request.get_json()
            
            # Store old values for logging
            old_values = requestor.to_dict()
            
            for key, value in data.items():
                if hasattr(requestor, key):
                    setattr(requestor, key, value)
            
            db.session.commit()
            
            # Log activity
            log_activity(
                action='UPDATE',
                entity_type='requestors',
                entity_id=requestor.id,
                entity_name=get_entity_name('requestors', data),
                old_values=old_values,
                new_values=data
            )
            
            return jsonify(requestor.to_dict())
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erreur lors de la mise à jour', 'error': str(e)}), 400

    @app.route("/api/requestors/<requestor_id>", methods=['DELETE'])
    def delete_requestor(requestor_id):
        try:
            requestor = Requestor.query.get(requestor_id)
            if not requestor:
                return jsonify({'message': 'Demandeur non trouvé'}), 404
            
            # Store values for logging before deletion
            old_values = requestor.to_dict()
            entity_name = get_entity_name('requestors', old_values)
            
            db.session.delete(requestor)
            db.session.commit()
            
            # Log activity
            log_activity(
                action='DELETE',
                entity_type='requestors',
                entity_id=requestor_id,
                entity_name=entity_name,
                old_values=old_values
            )
            
            return '', 204
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erreur lors de la suppression'}), 500

    @app.route("/api/requestors/bulk-delete", methods=['POST'])
    def bulk_delete_requestors():
        try:
            data = request.get_json()
            requestor_ids = data.get('ids', [])
            
            if not requestor_ids:
                return jsonify({'message': 'Aucun ID fourni'}), 400
            
            # Find requestors to delete
            requestors = Requestor.query.filter(Requestor.id.in_(requestor_ids)).all()
            deleted_count = len(requestors)
            
            # Delete requestors
            for requestor in requestors:
                db.session.delete(requestor)
            
            db.session.commit()
            
            # Log activity
            log_activity(
                action='BULK_DELETE',
                entity_type='requestors',
                entity_name=f'{deleted_count} demandeurs',
                old_values={'count': deleted_count, 'ids': requestor_ids}
            )
            
            return jsonify({
                'message': f'{deleted_count} demandeurs supprimés avec succès',
                'deletedCount': deleted_count
            }), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Bulk delete requestors error: {str(e)}")
            return jsonify({'message': 'Erreur lors de la suppression en masse'}), 500

    # Suppliers Import/Export routes
    @app.route("/api/suppliers/export", methods=['GET'])
    def export_suppliers():
        try:
            suppliers = Supplier.query.all()
            
            # Create DataFrame with supplier data
            data = []
            for supplier in suppliers:
                data.append({
                    'Nom': supplier.nom,
                    'Contact': supplier.contact or '',
                    'Téléphone': supplier.telephone or '',
                    'Email': supplier.email or '',
                    'Adresse': supplier.adresse or '',
                    'Conditions de paiement': supplier.conditions_paiement or '',
                    'Délai de livraison (jours)': supplier.delai_livraison or ''
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel file in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Fournisseurs', index=False)
            
            output.seek(0)
            
            # Create response
            filename = f"fournisseurs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            return response
            
        except Exception as e:
            logger.error(f"Suppliers export error: {str(e)}")
            return jsonify({'message': f'Erreur lors de l\'export: {str(e)}'}), 500

    @app.route("/api/suppliers/import", methods=['POST'])
    def import_suppliers():
        try:
            if 'file' not in request.files:
                return jsonify({'message': 'Aucun fichier fourni'}), 400
            
            file = request.files['file']
            if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
                return jsonify({'message': 'Fichier Excel requis (.xlsx ou .xls)'}), 400
            
            # Read Excel file
            df = pd.read_excel(file)
            
            # Expected columns
            expected_columns = ['Nom', 'Contact', 'Téléphone', 'Email', 'Adresse', 'Conditions de paiement', 'Délai de livraison (jours)']
            
            # Check if required column exists
            if 'Nom' not in df.columns:
                return jsonify({'message': 'Colonne "Nom" requise dans le fichier Excel'}), 400
            
            imported_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Check if supplier name is provided
                    if pd.isna(row.get('Nom')) or str(row.get('Nom')).strip() == '':
                        errors.append(f"Ligne {index + 2}: Nom requis")
                        continue
                    
                    # Check if supplier already exists
                    existing = Supplier.query.filter_by(nom=str(row['Nom']).strip()).first()
                    if existing:
                        errors.append(f"Ligne {index + 2}: Fournisseur '{row['Nom']}' existe déjà")
                        continue
                    
                    # Create new supplier
                    supplier = Supplier(
                        nom=str(row['Nom']).strip(),
                        contact=str(row.get('Contact', '')).strip() if not pd.isna(row.get('Contact')) else None,
                        telephone=str(row.get('Téléphone', '')).strip() if not pd.isna(row.get('Téléphone')) else None,
                        email=str(row.get('Email', '')).strip() if not pd.isna(row.get('Email')) else None,
                        adresse=str(row.get('Adresse', '')).strip() if not pd.isna(row.get('Adresse')) else None,
                        conditions_paiement=str(row.get('Conditions de paiement', '')).strip() if not pd.isna(row.get('Conditions de paiement')) else None,
                        delai_livraison=int(row.get('Délai de livraison (jours)', 0)) if not pd.isna(row.get('Délai de livraison (jours)')) and str(row.get('Délai de livraison (jours)')).strip() != '' else None
                    )
                    
                    db.session.add(supplier)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            db.session.commit()
            
            # Log activity
            if imported_count > 0:
                log_activity(
                    action='IMPORT',
                    entity_type='suppliers',
                    entity_name=f'{imported_count} fournisseurs'
                )
            
            return jsonify({
                'message': f'{imported_count} fournisseur(s) importé(s) avec succès',
                'imported': imported_count,
                'errors': errors
            })
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Suppliers import error: {str(e)}")
            return jsonify({'message': f'Erreur lors de l\'import: {str(e)}'}), 500

    # Requestors Import/Export routes
    @app.route("/api/requestors/export", methods=['GET'])
    def export_requestors():
        try:
            requestors = Requestor.query.all()
            
            # Create DataFrame with requestor data
            data = []
            for requestor in requestors:
                data.append({
                    'Prénom': requestor.prenom,
                    'Nom': requestor.nom,
                    'Département': requestor.departement,
                    'Poste': requestor.poste or '',
                    'Email': requestor.email or '',
                    'Téléphone': requestor.telephone or ''
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel file in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Demandeurs', index=False)
            
            output.seek(0)
            
            # Log activity
            log_activity(
                action='EXPORT',
                entity_type='requestors',
                entity_name=f'{len(requestors)} demandeurs'
            )
            
            # Create response
            filename = f"demandeurs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            return response
            
        except Exception as e:
            logger.error(f"Requestors export error: {str(e)}")
            return jsonify({'message': f'Erreur lors de l\'export: {str(e)}'}), 500

    @app.route("/api/requestors/import", methods=['POST'])
    def import_requestors():
        try:
            if 'file' not in request.files:
                return jsonify({'message': 'Aucun fichier fourni'}), 400
            
            file = request.files['file']
            if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
                return jsonify({'message': 'Fichier Excel requis (.xlsx ou .xls)'}), 400
            
            # Read Excel file
            df = pd.read_excel(file)
            
            # Check required columns
            required_columns = ['Prénom', 'Nom', 'Département']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return jsonify({'message': f'Colonnes requises manquantes: {", ".join(missing_columns)}'}), 400
            
            imported_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Check required fields
                    prenom = str(row.get('Prénom', '')).strip() if not pd.isna(row.get('Prénom')) else ''
                    nom = str(row.get('Nom', '')).strip() if not pd.isna(row.get('Nom')) else ''
                    departement = str(row.get('Département', '')).strip() if not pd.isna(row.get('Département')) else ''
                    
                    if not prenom or not nom or not departement:
                        errors.append(f"Ligne {index + 2}: Prénom, Nom et Département requis")
                        continue
                    
                    # Check if requestor already exists
                    existing = Requestor.query.filter_by(nom=nom, prenom=prenom, departement=departement).first()
                    if existing:
                        errors.append(f"Ligne {index + 2}: Demandeur '{prenom} {nom}' du département '{departement}' existe déjà")
                        continue
                    
                    # Create new requestor
                    requestor = Requestor(
                        prenom=prenom,
                        nom=nom,
                        departement=departement,
                        poste=str(row.get('Poste', '')).strip() if not pd.isna(row.get('Poste')) else None,
                        email=str(row.get('Email', '')).strip() if not pd.isna(row.get('Email')) else None,
                        telephone=str(row.get('Téléphone', '')).strip() if not pd.isna(row.get('Téléphone')) else None
                    )
                    
                    db.session.add(requestor)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            db.session.commit()
            
            # Log activity
            if imported_count > 0:
                log_activity(
                    action='IMPORT',
                    entity_type='requestors',
                    entity_name=f'{imported_count} demandeurs'
                )
            
            return jsonify({
                'message': f'{imported_count} demandeur(s) importé(s) avec succès',
                'imported': imported_count,
                'errors': errors
            })
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Requestors import error: {str(e)}")
            return jsonify({'message': f'Erreur lors de l\'import: {str(e)}'}), 500

    # Purchase Requests routes
    @app.route("/api/purchase-requests", methods=['GET'])
    def get_purchase_requests():
        try:
            requests = PurchaseRequest.query.all()
            return jsonify([request.to_dict() for request in requests])
        except Exception as e:
            print(f"Purchase requests error: {str(e)}")
            return jsonify({'message': f'Erreur lors de la récupération des demandes d\'achat: {str(e)}'}), 500

    @app.route("/api/purchase-requests", methods=['POST'])
    def create_purchase_request():
        try:
            data = request.get_json()
            
            # Generate request number
            count = PurchaseRequest.query.count() + 1
            numero_demande = f"DA-{datetime.utcnow().strftime('%Y')}-{count:04d}"
            
            # Calculate total
            total_estime = sum(item['quantiteDemandee'] * item['prixUnitaireEstime'] 
                             for item in data.get('articles', []))
            
            purchase_request = PurchaseRequest(
                numero_demande=numero_demande,
                date_demande=datetime.fromisoformat(data['dateDemande'].replace('Z', '+00:00')) if 'dateDemande' in data else datetime.utcnow(),
                requestor_id=data['requestorId'],
                observations=data.get('observations'),
                statut=data.get('statut', 'en_attente'),
                total_articles=len(data.get('articles', [])),
                total_estime=total_estime
            )
            db.session.add(purchase_request)
            db.session.flush()  # Get the ID
            
            # Add articles
            for article_data in data.get('articles', []):
                sous_total = article_data['quantiteDemandee'] * article_data['prixUnitaireEstime']
                
                item = PurchaseRequestItem(
                    purchase_request_id=purchase_request.id,
                    article_id=article_data['articleId'],
                    supplier_id=article_data.get('supplierId'),
                    quantite_demandee=article_data['quantiteDemandee'],
                    prix_unitaire_estime=article_data['prixUnitaireEstime'],
                    sous_total=sous_total
                )
                db.session.add(item)
            
            db.session.commit()
            return jsonify(purchase_request.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Données invalides', 'error': str(e)}), 400

    @app.route("/api/purchase-requests/<request_id>", methods=['PUT'])
    def update_purchase_request(request_id):
        try:
            purchase_request = PurchaseRequest.query.get(request_id)
            if not purchase_request:
                return jsonify({'message': 'Demande non trouvée'}), 404
            
            data = request.get_json()
            
            # Handle status updates
            if 'statut' in data:
                purchase_request.statut = data['statut']
            
            if 'observations' in data:
                purchase_request.observations = data['observations']
            
            # If updating articles, delete existing and recreate
            if 'articles' in data:
                # Delete existing items
                PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).delete()
                
                # Add new articles
                total_estime = 0
                for article_data in data['articles']:
                    sous_total = article_data['quantiteDemandee'] * article_data['prixUnitaireEstime']
                    total_estime += sous_total
                    
                    item = PurchaseRequestItem(
                        purchase_request_id=request_id,
                        article_id=article_data['articleId'],
                        supplier_id=article_data.get('supplierId'),
                        quantite_demandee=article_data['quantiteDemandee'],
                        prix_unitaire_estime=article_data['prixUnitaireEstime'],
                        sous_total=sous_total
                    )
                    db.session.add(item)
                
                purchase_request.total_articles = len(data['articles'])
                purchase_request.total_estime = total_estime
            
            db.session.commit()
            return jsonify(purchase_request.to_dict())
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erreur lors de la mise à jour', 'error': str(e)}), 400

    @app.route("/api/purchase-requests/<request_id>", methods=['DELETE'])
    def delete_purchase_request(request_id):
        try:
            purchase_request = PurchaseRequest.query.get(request_id)
            if not purchase_request:
                return jsonify({'message': 'Demande non trouvée'}), 404
            
            # Also delete related items
            PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).delete()
            db.session.delete(purchase_request)
            db.session.commit()
            return '', 204
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erreur lors de la suppression'}), 500

    # Purchase Request Items routes
    @app.route("/api/purchase-request-items", methods=['POST'])
    def create_purchase_request_item():
        try:
            data = request.get_json()
            item = PurchaseRequestItem(
                purchase_request_id=data['purchaseRequestId'],
                article_id=data['articleId'],
                quantite_demandee=data['quantiteDemandee'],
                supplier_id=data.get('supplierId'),
                prix_unitaire_estime=data.get('prixUnitaireEstime'),
                observations=data.get('observations')
            )
            db.session.add(item)
            
            # Update total articles count
            item_count = PurchaseRequestItem.query.filter_by(purchase_request_id=data['purchaseRequestId']).count() + 1
            purchase_request = PurchaseRequest.query.get(data['purchaseRequestId'])
            if purchase_request:
                purchase_request.total_articles = item_count
            
            db.session.commit()
            return jsonify(item.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Données invalides', 'error': str(e)}), 400

    @app.route("/api/purchase-requests/<request_id>/items", methods=['GET'])
    def get_purchase_request_items(request_id):
        try:
            items = PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).all()
            return jsonify([item.to_dict() for item in items])
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération des éléments'}), 500
    
    # Convert Purchase Request to Reception
    @app.route("/api/purchase-requests/<request_id>/convert-reception", methods=['POST'])
    def convert_purchase_request_to_reception(request_id):
        try:
            data = request.get_json()
            
            # Get purchase request and items
            purchase_request = PurchaseRequest.query.get_or_404(request_id)
            items = PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).all()
            
            # Create reception records for each article
            receptions_created = []
            
            for reception_article in data.get('articles', []):
                # Find corresponding item
                item = next((i for i in items if i.id == reception_article['itemId']), None)
                if not item:
                    continue
                
                # Get supplier_id from form data first, then from item, then from article's default supplier
                supplier_id = reception_article.get('supplierId') or item.supplier_id
                if not supplier_id:
                    # Fallback: get from article's default supplier
                    article = Article.query.get(item.article_id)
                    if article and article.fournisseur_id:
                        supplier_id = article.fournisseur_id
                
                # If no supplier found, create with null supplier (will be updated later)
                if not supplier_id:
                    # Try to get the default supplier from the first available supplier
                    default_supplier = Supplier.query.first()
                    if default_supplier:
                        supplier_id = default_supplier.id
                    else:
                        return jsonify({'message': f'Aucun fournisseur disponible dans le système pour l\'article {item.article.designation if item.article else "inconnu"}'}), 400
                
                # Create reception record
                reception = Reception(
                    date_reception=datetime.strptime(data['dateReception'], '%Y-%m-%d').date(),
                    supplier_id=supplier_id,
                    article_id=item.article_id,
                    quantite_recue=reception_article['quantiteRecue'],
                    prix_unitaire=reception_article['prixUnitaire'],
                    numero_bon_livraison=data.get('numeroBonLivraison'),
                    observations=data.get('observations', f"Réception pour demande d'achat #{purchase_request.id[:8]}")
                )
                db.session.add(reception)
                
                # Update article stock
                article = Article.query.get(item.article_id)
                if article:
                    article.stock_actuel += reception_article['quantiteRecue']
                
                receptions_created.append(reception)
            
            db.session.commit()
            
            return jsonify({
                'message': f'{len(receptions_created)} réceptions créées avec succès',
                'receptions': [r.to_dict() for r in receptions_created]
            }), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': f'Erreur lors de la conversion: {str(e)}'}), 500

    # Complete purchase request creation
    @app.route("/api/purchase-requests/complete", methods=['POST'])
    def create_complete_purchase_request():
        try:
            data = request.get_json()
            
            # Create purchase request header
            purchase_request = PurchaseRequest(
                date_demande=datetime.fromisoformat(data['dateDemande'].replace('Z', '+00:00')) if 'dateDemande' in data else datetime.utcnow(),
                requestor_id=data['requestorId'],
                observations=data.get('observations'),
                total_articles=len(data['items']),
                statut='en_attente'
            )
            db.session.add(purchase_request)
            db.session.flush()  # Get the ID
            
            # Create items
            items = []
            for item_data in data['items']:
                item = PurchaseRequestItem(
                    purchase_request_id=purchase_request.id,
                    article_id=item_data['articleId'],
                    quantite_demandee=item_data['quantiteDemandee'],
                    supplier_id=item_data.get('supplierId'),
                    prix_unitaire_estime=item_data.get('prixUnitaireEstime'),
                    observations=item_data.get('observations')
                )
                db.session.add(item)
                items.append(item)
            
            db.session.commit()
            
            return jsonify({
                **purchase_request.to_dict(),
                'items': [item.to_dict() for item in items]
            }), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Données invalides', 'error': str(e)}), 400

    # Reception routes
    @app.route("/api/receptions", methods=['GET'])
    def get_receptions():
        try:
            receptions = Reception.query.all()
            return jsonify([reception.to_dict() for reception in receptions])
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération des réceptions'}), 500

    @app.route("/api/receptions", methods=['POST'])
    def create_reception():
        try:
            data = request.get_json()
            reception = Reception(
                date_reception=datetime.fromisoformat(data['dateReception'].replace('Z', '+00:00')) if 'dateReception' in data else datetime.utcnow(),
                supplier_id=data['supplierId'],
                article_id=data['articleId'],
                quantite_recue=data['quantiteRecue'],
                prix_unitaire=data.get('prixUnitaire'),
                numero_bon_livraison=data.get('numeroBonLivraison'),
                observations=data.get('observations')
            )
            db.session.add(reception)
            
            # Update article stock
            article = Article.query.get(data['articleId'])
            if article:
                article.stock_actuel += data['quantiteRecue']
            
            db.session.commit()
            return jsonify(reception.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Données invalides', 'error': str(e)}), 400

    # Outbound routes
    @app.route("/api/outbounds", methods=['GET'])
    def get_outbounds():
        try:
            outbounds = Outbound.query.all()
            return jsonify([outbound.to_dict() for outbound in outbounds])
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération des sorties'}), 500

    @app.route("/api/outbounds", methods=['POST'])
    def create_outbound():
        try:
            data = request.get_json()
            
            # Generate a unique transaction number
            numero_sortie = f"OUT-{datetime.now().strftime('%Y%m%d')}-{str(int(time.time()))[-6:]}"
            
            # Handle multi-article outbound transaction
            articles_data = data.get('articles', [])
            if not articles_data:
                return jsonify({'message': 'Aucun article spécifié'}), 400
            
            created_outbounds = []
            
            for article_item in articles_data:
                outbound = Outbound(
                    numero_sortie=numero_sortie,
                    date_sortie=datetime.now(),
                    requestor_id=data.get('requestorId'),
                    article_id=article_item['articleId'],
                    quantite_sortie=article_item['quantiteSortie'],
                    motif_sortie=data['motifSortie'],
                    observations=data.get('observations', '')
                )
                
                db.session.add(outbound)
                created_outbounds.append(outbound)
                
                # Update article stock
                article = Article.query.get(article_item['articleId'])
                if article:
                    article.stock_actuel -= article_item['quantiteSortie']
                    if article.stock_actuel < 0:
                        article.stock_actuel = 0
            
            db.session.commit()
            return jsonify({
                'message': 'Sortie créée avec succès',
                'numeroSortie': numero_sortie,
                'outbounds': [outbound.to_dict() for outbound in created_outbounds]
            }), 201
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': f'Erreur lors de la création de la sortie: {str(e)}'}), 500

    # Analytics routes
    @app.route("/api/analytics/overview", methods=['GET'])
    def get_analytics_overview():
        try:
            total_articles = Article.query.count()
            total_suppliers = Supplier.query.count()
            total_requests = PurchaseRequest.query.count()
            low_stock_count = Article.query.filter(Article.stock_actuel <= Article.seuil_minimum).count()
            
            return jsonify({
                'totalArticles': total_articles,
                'totalSuppliers': total_suppliers,
                'totalRequests': total_requests,
                'lowStockCount': low_stock_count
            })
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération des analytics'}), 500

    # Dashboard stats endpoint  
    @app.route("/api/dashboard/stats", methods=['GET'])
    def get_dashboard_stats():
        try:
            # Basic counts
            total_articles = Article.query.count()
            total_suppliers = Supplier.query.count()
            total_requestors = Requestor.query.count()
            total_requests = PurchaseRequest.query.count()
            total_receptions = Reception.query.count()
            total_outbounds = Outbound.query.count()
            
            # Low stock count (articles at or below minimum threshold)
            low_stock_count = Article.query.filter(Article.stock_actuel <= Article.seuil_minimum).count()
            
            # Pending purchase requests
            pending_requests = PurchaseRequest.query.filter(PurchaseRequest.statut == 'en_attente').count()
            
            # Calculate total stock value
            articles = Article.query.all()
            stock_value = 0
            for article in articles:
                if article.stock_actuel > 0 and article.prix_unitaire:
                    stock_value += article.stock_actuel * article.prix_unitaire
            
            # Get purchase request status distribution for charts
            status_counts = {
                'en_attente': PurchaseRequest.query.filter(PurchaseRequest.statut == 'en_attente').count(),
                'approuve': PurchaseRequest.query.filter(PurchaseRequest.statut == 'approuve').count(),
                'refuse': PurchaseRequest.query.filter(PurchaseRequest.statut == 'refuse').count(),
                'commande': PurchaseRequest.query.filter(PurchaseRequest.statut == 'commande').count(),
                'recu': PurchaseRequest.query.filter(PurchaseRequest.statut == 'recu').count()
            }
            
            # Get recent activity
            recent_receptions = Reception.query.order_by(desc(Reception.date_reception)).limit(5).all()
            recent_outbounds = Outbound.query.order_by(desc(Outbound.date_sortie)).limit(5).all()
            
            return jsonify({
                'totalArticles': total_articles,
                'totalSuppliers': total_suppliers,
                'totalRequestors': total_requestors,
                'totalRequests': total_requests,
                'totalReceptions': total_receptions,
                'totalOutbounds': total_outbounds,
                'lowStock': low_stock_count,
                'pendingRequests': pending_requests,
                'stockValue': stock_value,
                'statusCounts': status_counts,
                'recentReceptions': [reception.to_dict() for reception in recent_receptions],
                'recentOutbounds': [outbound.to_dict() for outbound in recent_outbounds]
            })
        except Exception as e:
            print(f"Dashboard stats error: {str(e)}")
            return jsonify({'message': f'Erreur lors de la récupération des statistiques: {str(e)}'}), 500

    # Stock status analytics
    @app.route("/api/stock-status/analytics", methods=['GET'])
    def get_stock_status_analytics():
        try:
            articles = Article.query.all()
            
            # Categorize by stock levels
            critical = []  # 0 stock
            low = []       # below minimum
            medium = []    # 1-2x minimum
            good = []      # above 2x minimum
            
            for article in articles:
                seuil = article.seuil_minimum or 10
                if article.stock_actuel == 0:
                    critical.append(article.to_dict())
                elif article.stock_actuel <= seuil:
                    low.append(article.to_dict())
                elif article.stock_actuel <= seuil * 2:
                    medium.append(article.to_dict())
                else:
                    good.append(article.to_dict())
            
            return jsonify({
                'critical': critical,
                'low': low,
                'medium': medium,
                'good': good,
                'summary': {
                    'critical_count': len(critical),
                    'low_count': len(low),
                    'medium_count': len(medium),
                    'good_count': len(good),
                    'total_count': len(articles)
                }
            })
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération du statut stock'}), 500

    # Purchase follow-up analytics
    @app.route("/api/purchase-follow/status", methods=['GET'])
    def get_purchase_follow_status():
        try:
            pending = PurchaseRequest.query.filter_by(statut='en_attente').all()
            approved = PurchaseRequest.query.filter_by(statut='approuve').all()
            ordered = PurchaseRequest.query.filter_by(statut='commande').all()
            refused = PurchaseRequest.query.filter_by(statut='refuse').all()
            
            return jsonify({
                'pending': [req.to_dict() for req in pending],
                'approved': [req.to_dict() for req in approved],
                'ordered': [req.to_dict() for req in ordered],
                'refused': [req.to_dict() for req in refused]
            })
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération du suivi'}), 500

    # Reports endpoints
    @app.route("/api/reports/stock", methods=['GET'])
    def generate_stock_report():
        try:
            articles = Article.query.all()
            report_data = {
                'timestamp': datetime.utcnow().isoformat(),
                'total_articles': len(articles),
                'articles': [article.to_dict() for article in articles],
                'summary': {
                    'total_value': sum(article.prix_unitaire * article.stock_actuel for article in articles if article.prix_unitaire),
                    'low_stock_count': len([a for a in articles if a.stock_actuel <= (a.seuil_minimum or 10)])
                }
            }
            return jsonify(report_data)
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la génération du rapport'}), 500

    # Settings endpoints
    @app.route("/api/settings", methods=['GET'])
    def get_settings():
        try:
            # Load settings from file or database
            import json
            import os
            settings_file = 'settings.json'
            
            if os.path.exists(settings_file):
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            else:
                # Default settings
                settings = {
                    'currency': 'MAD',
                    'language': 'fr',
                    'dateFormat': 'dd/mm/yyyy',
                    'theme': 'light',
                    'notifications': {
                        'stock': True,
                        'requests': True,
                        'deliveries': True,
                        'reports': False
                    },
                    'paginationLimit': 50,
                    'autoRefresh': True,
                    'security': {
                        'passwordLength': 8,
                        'passwordExpiry': 90,
                        'requireUppercase': True,
                        'requireNumbers': True,
                        'requireSpecial': False,
                        'sessionTimeout': 30,
                        'autoLogout': True,
                        'logLogins': True,
                        'logDataChanges': True,
                        'logExports': False
                    },
                    'backup': {
                        'frequency': 'daily',
                        'time': '02:00',
                        'retention': 7
                    },
                    'integrations': {
                        'apiEnabled': False,
                        'scannerEnabled': False,
                        'scannerType': 'camera'
                    },
                    'advanced': {
                        'dbPoolSize': 10,
                        'dbTimeout': 30,
                        'dbAutoVacuum': False,
                        'debugMode': False,
                        'smartCache': True,
                        'strictValidation': False,
                        'maintenanceMode': False,
                        'searchLimit': 500,
                        'cacheDuration': 15,
                        'batchSize': 100,
                        'metricsCpu': True,
                        'metricsMemory': True,
                        'metricsDisk': False,
                        'alertSlowQueries': False,
                        'alertHighCpu': False,
                        'alertErrors': False
                    },
                    'customization': {
                        'primaryColor': '#003d9d',
                        'secondaryColor': '#6B7280',
                        'navStyle': 'top',
                        'displayDensity': 'comfortable',
                        'roundedCorners': True,
                        'companyName': 'StockCéramique',
                        'companyTagline': '',
                        'moduleAnalytics': True,
                        'moduleQr': True,
                        'moduleExport': True,
                        'moduleWorkflow': False,
                        'moduleSmartAlerts': False,
                        'moduleOffline': False,
                        'widgetStats': True,
                        'widgetCharts': True,
                        'widgetAlerts': True,
                        'widgetRecent': True,
                        'widgetRequests': False,
                        'dashboardLayout': 'masonry'
                    }
                }
                
            return jsonify(settings)
        except Exception as e:
            return jsonify({'message': 'Erreur lors du chargement des paramètres'}), 500
    
    @app.route("/api/settings", methods=['POST'])
    def save_settings():
        try:
            import json
            settings = request.get_json()
            
            # Save settings to file
            with open('settings.json', 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=2, ensure_ascii=False)
                
            return jsonify({'message': 'Paramètres sauvegardés avec succès'})
        except Exception as e:
            return jsonify({'message': f'Erreur lors de la sauvegarde: {str(e)}'}), 500
    
    @app.route("/api/settings/backup", methods=['POST'])
    def create_backup():
        try:
            import json
            import shutil
            from datetime import datetime
            
            # Create backup of database and settings
            backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            backup_dir = f"backups/{backup_name}"
            
            import os
            os.makedirs(backup_dir, exist_ok=True)
            
            # Copy database file
            if os.path.exists('instance/stockceramique.db'):
                shutil.copy2('instance/stockceramique.db', f"{backup_dir}/database.db")
            
            # Copy settings
            if os.path.exists('settings.json'):
                shutil.copy2('settings.json', f"{backup_dir}/settings.json")
                
            # Create backup info file
            backup_info = {
                'name': backup_name,
                'created': datetime.now().isoformat(),
                'type': 'manual',
                'size': 0  # Would calculate actual size
            }
            
            with open(f"{backup_dir}/info.json", 'w') as f:
                json.dump(backup_info, f, indent=2)
                
            return jsonify({
                'message': 'Sauvegarde créée avec succès',
                'backup': backup_info
            })
        except Exception as e:
            return jsonify({'message': f'Erreur lors de la création de sauvegarde: {str(e)}'}), 500
    
    @app.route("/api/settings/system-info", methods=['GET'])
    def get_system_info():
        try:
            import psutil
            import os
            from datetime import datetime
            
            # Get system metrics
            system_info = {
                'cpu': {
                    'usage': psutil.cpu_percent(interval=1),
                    'cores': psutil.cpu_count()
                },
                'memory': {
                    'total': psutil.virtual_memory().total,
                    'used': psutil.virtual_memory().used,
                    'percent': psutil.virtual_memory().percent
                },
                'disk': {
                    'total': psutil.disk_usage('.').total,
                    'used': psutil.disk_usage('.').used,
                    'percent': (psutil.disk_usage('.').used / psutil.disk_usage('.').total) * 100
                },
                'database': {
                    'size': os.path.getsize('instance/stockceramique.db') if os.path.exists('instance/stockceramique.db') else 0,
                    'articles': Article.query.count(),
                    'suppliers': Supplier.query.count(),
                    'requests': PurchaseRequest.query.count()
                },
                'uptime': datetime.now().isoformat(),
                'version': '1.0.0'
            }
            
            return jsonify(system_info)
        except ImportError:
            # If psutil is not available, return basic info
            return jsonify({
                'message': 'Informations système limitées',
                'database': {
                    'articles': Article.query.count(),
                    'suppliers': Supplier.query.count(),
                    'requests': PurchaseRequest.query.count()
                },
                'version': '1.0.0'
            })
        except Exception as e:
            return jsonify({'message': f'Erreur lors de la récupération des informations système: {str(e)}'}), 500

    @app.route("/api/settings/categories", methods=['GET'])
    def get_categories():
        try:
            categories = db.session.query(Article.categorie).distinct().all()
            return jsonify([cat[0] for cat in categories if cat[0]])
        except Exception as e:
            return jsonify({'message': 'Erreur lors de la récupération des catégories'}), 500
            
    @app.route("/api/settings/categories", methods=['POST'])
    def add_category():
        try:
            data = request.get_json()
            category_name = data.get('name')
            
            if not category_name:
                return jsonify({'message': 'Nom de catégorie requis'}), 400
                
            # Check if category already exists
            existing = db.session.query(Article.categorie).filter_by(categorie=category_name).first()
            if existing:
                return jsonify({'message': 'Cette catégorie existe déjà'}), 400
                
            return jsonify({'message': 'Catégorie ajoutée avec succès'})
        except Exception as e:
            return jsonify({'message': f'Erreur lors de l\'ajout de catégorie: {str(e)}'}), 500

    @app.route("/api/settings/units", methods=['POST'])
    def add_unit():
        try:
            data = request.get_json()
            code = data.get('code')
            description = data.get('description')
            
            if not code or not description:
                return jsonify({'message': 'Code et description requis'}), 400
                
            # This would typically save to a units configuration table
            return jsonify({'message': 'Unité ajoutée avec succès'})
        except Exception as e:
            return jsonify({'message': f'Erreur lors de l\'ajout d\'unité: {str(e)}'}), 500
    
    @app.route("/api/settings/departments", methods=['POST'])
    def add_department():
        try:
            data = request.get_json()
            department_name = data.get('name')
            
            if not department_name:
                return jsonify({'message': 'Nom de département requis'}), 400
                
            # This would save to departments table or update requestors config
            return jsonify({'message': 'Département ajouté avec succès'})
        except Exception as e:
            return jsonify({'message': f'Erreur lors de l\'ajout de département: {str(e)}'}), 500
    
    @app.route("/api/profile", methods=['GET'])
    def get_profile():
        try:
            # Mock user profile data - in real app would get from session/auth
            profile = {
                'firstName': 'Admin',
                'lastName': 'User',
                'email': 'admin@stockceramique.com',
                'phone': '+212 123 456 789',
                'position': 'Gestionnaire Inventaire',
                'department': 'administration',
                'address': '123 Rue de l\'Industrie',
                'city': 'Casablanca',
                'postalCode': '20000',
                'avatar': None,
                'lastLogin': '2024-08-23T09:15:00',
                'preferences': {
                    'language': 'fr',
                    'timezone': 'Africa/Casablanca',
                    'defaultPage': '/',
                    'pagination': 50,
                    'notifications': {
                        'email': True,
                        'push': False,
                        'weekly': True
                    }
                },
                'statistics': {
                    'requestsCreated': 127,
                    'articlesAdded': 89,
                    'receptionsProcessed': 45,
                    'loginCount': 156
                }
            }
            return jsonify(profile)
        except Exception as e:
            return jsonify({'message': f'Erreur lors du chargement du profil: {str(e)}'}), 500
    
    @app.route("/api/profile", methods=['POST'])
    def save_profile():
        try:
            profile_data = request.get_json()
            # In real app, would save to user database
            return jsonify({'message': 'Profil sauvegardé avec succès'})
        except Exception as e:
            return jsonify({'message': f'Erreur lors de la sauvegarde du profil: {str(e)}'}), 500

    # Bulk Import/Export Articles
    @app.route("/api/articles/export", methods=['GET'])
    def export_articles():
        try:
            # Get export format from query parameters
            format_type = request.args.get('format', 'csv')
            
            # Get all articles
            articles = Article.query.all()
            
            # Convert to list of dictionaries
            data = []
            for article in articles:
                data.append({
                    'Code Article': article.code_article,
                    'Désignation': article.designation,
                    'Catégorie': article.categorie,
                    'Marque': article.marque or '',
                    'Référence': article.reference or '',
                    'Stock Initial': article.stock_initial,
                    'Stock Actuel': article.stock_actuel,
                    'Unité': article.unite,
                    'Prix Unitaire': float(article.prix_unitaire) if article.prix_unitaire else 0,
                    'Seuil Minimum': article.seuil_minimum,
                    'Fournisseur ID': article.fournisseur_id or '',
                    'Date Création': article.created_at.strftime('%Y-%m-%d %H:%M:%S') if article.created_at else ''
                })
            
            # Create dataframe
            df = pd.DataFrame(data)
            
            if format_type == 'excel':
                # Export to Excel
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Articles', index=False)
                output.seek(0)
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename=articles_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                return response
            else:
                # Export to CSV
                output = io.StringIO()
                df.to_csv(output, index=False, encoding='utf-8')
                output.seek(0)
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'text/csv; charset=utf-8'
                response.headers['Content-Disposition'] = f'attachment; filename=articles_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
                return response
                
        except Exception as e:
            return jsonify({'message': f'Erreur lors de l\'export: {str(e)}'}), 500

    # Modern export endpoints with options
    @app.route("/api/articles/export/pdf", methods=['POST'])
    def export_articles_pdf():
        try:
            # Get export options from request body
            options = request.get_json() or {}
            include_stock = options.get('includeStock', True)
            include_prices = options.get('includePrices', True)
            include_suppliers = options.get('includeSuppliers', True)
            
            # Get all articles
            articles = Article.query.all()
            
            # Prepare data based on options
            data = []
            for article in articles:
                row = {
                    'Code Article': article.code_article,
                    'Désignation': article.designation,
                    'Catégorie': article.categorie,
                    'Marque': article.marque or '',
                    'Référence': article.reference or '',
                    'Unité': article.unite,
                }
                
                if include_stock:
                    row.update({
                        'Stock Initial': article.stock_initial,
                        'Stock Actuel': article.stock_actuel,
                        'Seuil Minimum': article.seuil_minimum,
                    })
                
                if include_prices:
                    row['Prix Unitaire'] = f"{float(article.prix_unitaire):.2f} MAD" if article.prix_unitaire else "0.00 MAD"
                
                if include_suppliers:
                    row['Fournisseur ID'] = article.fournisseur_id or ''
                
                row['Date Création'] = article.created_at.strftime('%Y-%m-%d') if article.created_at else ''
                data.append(row)
            
            # Create HTML content for PDF
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Export Articles - StockCéramique</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    .header {{ text-align: center; margin-bottom: 30px; }}
                    .company {{ color: #003d9d; font-size: 24px; font-weight: bold; }}
                    .title {{ color: #666; font-size: 18px; margin-top: 10px; }}
                    .date {{ color: #999; font-size: 12px; margin-top: 5px; }}
                    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; font-size: 10px; }}
                    th {{ background-color: #003d9d; color: white; font-weight: bold; }}
                    tr:nth-child(even) {{ background-color: #f9f9f9; }}
                    .footer {{ margin-top: 30px; text-align: center; font-size: 10px; color: #666; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="company">StockCéramique</div>
                    <div class="title">Export Articles</div>
                    <div class="date">Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</div>
                </div>
                
                <table>
                    <thead>
                        <tr>
                            {''.join(f'<th>{col}</th>' for col in data[0].keys() if data)}
                        </tr>
                    </thead>
                    <tbody>
                        {''.join('<tr>' + ''.join(f'<td>{value}</td>' for value in row.values()) + '</tr>' for row in data)}
                    </tbody>
                </table>
                
                <div class="footer">
                    <p>Total: {len(data)} articles</p>
                    <p>StockCéramique - Système de Gestion d'Inventaire</p>
                </div>
            </body>
            </html>
            """
            
            # Generate actual PDF using weasyprint
            from weasyprint import HTML
            pdf_buffer = io.BytesIO()
            HTML(string=html_content).write_pdf(pdf_buffer)
            pdf_buffer.seek(0)
            
            response = make_response(pdf_buffer.getvalue())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=articles_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            return response
            
        except Exception as e:
            return jsonify({'message': f'Erreur lors de l\'export PDF: {str(e)}'}), 500

    @app.route("/api/articles/export/excel", methods=['POST'])
    def export_articles_excel():
        try:
            # Get export options from request body
            options = request.get_json() or {}
            include_stock = options.get('includeStock', True)
            include_prices = options.get('includePrices', True)
            include_suppliers = options.get('includeSuppliers', True)
            
            # Get all articles
            articles = Article.query.all()
            
            # Prepare data based on options
            data = []
            for article in articles:
                row = {
                    'Code Article': article.code_article,
                    'Désignation': article.designation,
                    'Catégorie': article.categorie,
                    'Marque': article.marque or '',
                    'Référence': article.reference or '',
                    'Unité': article.unite,
                }
                
                if include_stock:
                    row.update({
                        'Stock Initial': article.stock_initial,
                        'Stock Actuel': article.stock_actuel,
                        'Seuil Minimum': article.seuil_minimum,
                    })
                
                if include_prices:
                    row['Prix Unitaire'] = float(article.prix_unitaire) if article.prix_unitaire else 0
                
                if include_suppliers:
                    row['Fournisseur ID'] = article.fournisseur_id or ''
                
                row['Date Création'] = article.created_at.strftime('%Y-%m-%d %H:%M:%S') if article.created_at else ''
                data.append(row)
            
            # Create dataframe
            df = pd.DataFrame(data)
            
            # Export to Excel with styling
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Articles', index=False)
                
                # Get the workbook and worksheet for styling
                workbook = writer.book
                worksheet = writer.sheets['Articles']
                
                # Style the header row
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='003d9d', end_color='003d9d', fill_type='solid')
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=articles_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            return response
            
        except Exception as e:
            return jsonify({'message': f'Erreur lors de l\'export Excel: {str(e)}'}), 500

    @app.route("/api/articles/export/csv", methods=['POST'])
    def export_articles_csv():
        try:
            # Get export options from request body
            options = request.get_json() or {}
            include_stock = options.get('includeStock', True)
            include_prices = options.get('includePrices', True)
            include_suppliers = options.get('includeSuppliers', True)
            
            # Get all articles
            articles = Article.query.all()
            
            # Prepare data based on options
            data = []
            for article in articles:
                row = {
                    'Code Article': article.code_article,
                    'Désignation': article.designation,
                    'Catégorie': article.categorie,
                    'Marque': article.marque or '',
                    'Référence': article.reference or '',
                    'Unité': article.unite,
                }
                
                if include_stock:
                    row.update({
                        'Stock Initial': article.stock_initial,
                        'Stock Actuel': article.stock_actuel,
                        'Seuil Minimum': article.seuil_minimum,
                    })
                
                if include_prices:
                    row['Prix Unitaire'] = float(article.prix_unitaire) if article.prix_unitaire else 0
                
                if include_suppliers:
                    row['Fournisseur ID'] = article.fournisseur_id or ''
                
                row['Date Création'] = article.created_at.strftime('%Y-%m-%d %H:%M:%S') if article.created_at else ''
                data.append(row)
            
            # Create dataframe
            df = pd.DataFrame(data)
            
            # Export to CSV
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = f'attachment; filename=articles_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            return response
            
        except Exception as e:
            return jsonify({'message': f'Erreur lors de l\'export CSV: {str(e)}'}), 500

    @app.route("/api/articles/import", methods=['POST'])
    def import_articles():
        try:
            if 'file' not in request.files:
                return jsonify({'message': 'Aucun fichier fourni'}), 400
            
            file = request.files['file']
            if file.filename == '':
                return jsonify({'message': 'Nom de fichier vide'}), 400
            
            # Check file extension
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            
            if file_ext not in ['csv', 'xlsx', 'xls']:
                return jsonify({'message': 'Format de fichier non supporté. Utilisez CSV ou Excel.'}), 400
            
            # Read file content
            try:
                if file_ext == 'csv':
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)
            except Exception as e:
                return jsonify({'message': f'Erreur lors de la lecture du fichier: {str(e)}'}), 400
            
            # Expected columns (mapping from display names to database fields)
            expected_columns = {
                'Code Article': 'code_article',
                'Désignation': 'designation', 
                'Catégorie': 'categorie',
                'Marque': 'marque',
                'Référence': 'reference',
                'Stock Initial': 'stock_initial',
                'Stock Actuel': 'stock_actuel',
                'Unité': 'unite',
                'Prix Unitaire': 'prix_unitaire',
                'Seuil Minimum': 'seuil_minimum',
                'Fournisseur ID': 'fournisseur_id'
            }
            
            # Check required columns
            required_columns = ['Code Article', 'Désignation', 'Catégorie']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({'message': f'Colonnes manquantes: {", ".join(missing_columns)}'}), 400
            
            # Process rows
            created_count = 0
            updated_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Check if article exists
                    existing_article = Article.query.filter_by(code_article=row['Code Article']).first()
                    
                    if existing_article:
                        # Update existing article
                        for display_name, db_field in expected_columns.items():
                            if display_name in df.columns and pd.notna(row[display_name]):
                                value = row[display_name]
                                if db_field in ['stock_initial', 'stock_actuel', 'seuil_minimum']:
                                    value = int(value) if pd.notna(value) else 0
                                elif db_field == 'prix_unitaire':
                                    value = float(value) if pd.notna(value) else None
                                elif db_field in ['marque', 'reference', 'fournisseur_id']:
                                    value = str(value) if pd.notna(value) else None
                                else:
                                    value = str(value) if pd.notna(value) else ''
                                setattr(existing_article, db_field, value)
                        updated_count += 1
                    else:
                        # Create new article
                        article_data = {}
                        for display_name, db_field in expected_columns.items():
                            if display_name in df.columns:
                                value = row[display_name]
                                if db_field in ['stock_initial', 'stock_actuel', 'seuil_minimum']:
                                    article_data[db_field] = int(value) if pd.notna(value) else (10 if db_field == 'seuil_minimum' else 0)
                                elif db_field == 'prix_unitaire':
                                    article_data[db_field] = float(value) if pd.notna(value) else None
                                elif db_field in ['marque', 'reference', 'fournisseur_id']:
                                    article_data[db_field] = str(value) if pd.notna(value) else None
                                else:
                                    article_data[db_field] = str(value) if pd.notna(value) else ('pcs' if db_field == 'unite' else '')
                        
                        # Set defaults for required fields
                        if 'unite' not in article_data or not article_data['unite']:
                            article_data['unite'] = 'pcs'
                        if 'seuil_minimum' not in article_data:
                            article_data['seuil_minimum'] = 10
                        if 'stock_initial' not in article_data:
                            article_data['stock_initial'] = 0
                        if 'stock_actuel' not in article_data:
                            article_data['stock_actuel'] = article_data['stock_initial']
                        
                        article = Article(**article_data)
                        db.session.add(article)
                        created_count += 1
                        
                except Exception as e:
                    errors.append(f'Ligne {index + 2}: {str(e)}')
                    continue
            
            # Commit changes
            try:
                db.session.commit()
                message = f'Import terminé: {created_count} articles créés, {updated_count} articles mis à jour'
                if errors:
                    message += f'. Erreurs: {len(errors)} lignes ignorées'
                return jsonify({
                    'message': message,
                    'created': created_count,
                    'updated': updated_count,
                    'errors': errors[:5]  # Show first 5 errors only
                }), 200
            except Exception as e:
                db.session.rollback()
                return jsonify({'message': f'Erreur lors de la sauvegarde: {str(e)}'}), 500
                
        except Exception as e:
            return jsonify({'message': f'Erreur lors de l\'import: {str(e)}'}), 500

    @app.route("/api/articles/template", methods=['GET'])
    def get_import_template():
        try:
            # Create template with headers and example data
            template_data = [{
                'Code Article': 'ART001',
                'Désignation': 'Exemple Article',
                'Catégorie': 'Électronique',
                'Marque': 'Samsung',
                'Référence': 'REF001',
                'Stock Initial': 100,
                'Stock Actuel': 95,
                'Unité': 'pcs',
                'Prix Unitaire': 25.50,
                'Seuil Minimum': 10,
                'Fournisseur ID': '',
                'Date Création': ''
            }]
            
            df = pd.DataFrame(template_data)
            
            # Export as Excel template
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Template Articles', index=False)
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = 'attachment; filename=template_import_articles.xlsx'
            return response
            
        except Exception as e:
            return jsonify({'message': f'Erreur lors de la génération du template: {str(e)}'}), 500